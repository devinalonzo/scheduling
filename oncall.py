import os
import random
import datetime
import platform
import sys
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.borders import Border, Side
#import logging

# Install colorama if not already installed
try:
    from colorama import init, Fore, Back, Style
except ImportError:
    os.system('pip install colorama')
    from colorama import init, Fore, Back, Style

init(autoreset=True)

# Constants
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
SCHEDULING_FOLDER = os.path.join(DESKTOP, "On-Call Scheduling")
# SCHEDULE_FILE will be set dynamically based on the year
# Set up logging
LOG_FILE = os.path.join(SCHEDULING_FOLDER, 'schedule_log.txt')
#logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Ensure the folder exists
os.makedirs(SCHEDULING_FOLDER, exist_ok=True)

# Predefined employee list for 'devintest' mode
DEVINTEST_EMPLOYEES = [
    'Alice Smith', 'Bob Johnson', 'Carol Williams', 'David Jones',
    'Eve Brown', 'Frank Miller', 'Grace Wilson', 'Hank Moore',
    'Ivy Taylor', 'Jack Anderson', 'Kathy Thomas', 'Leo Jackson',
    'Mona White', 'Nate Harris', 'Olivia Martin', 'Paul Thompson',
    'Quincy Garcia', 'Rachel Martinez'
]

def clear_screen():
    if platform.system() == "Windows":
        os.system('cls')
    else:
        os.system('clear')

def center_text(text, width=None):
    if width is None:
        width = os.get_terminal_size().columns
    return text.center(width)

def print_centered(text, color=''):
    print(color + center_text(text))

def pause(message="Press Enter to continue..."):
    input(message)

def get_schedule_file(year):
    return os.path.join(SCHEDULING_FOLDER, f"{year} On Call Scheduling.xlsx")

def get_workbook(year):
    schedule_file = get_schedule_file(year)
    if os.path.exists(schedule_file):
        workbook = load_workbook(schedule_file)
    else:
        workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            sheet = workbook['Sheet']
            workbook.remove(sheet)
        # Create required sheets
        workbook.create_sheet('Employee List')
        workbook.create_sheet('Schedule Changes')
        workbook.create_sheet('Reports')
        workbook.create_sheet('Original Reports')
        # Create sheets for each month
        for month in range(1, 13):
            month_name = datetime.date(year, month, 1).strftime('%B')
            workbook.create_sheet(month_name)
        workbook.save(schedule_file)
    return workbook

def load_employees():
    # Assuming the employee list is the same across years
    # You can adjust this if needed
    year = datetime.datetime.now().year
    workbook = get_workbook(year)
    if 'Employee List' not in workbook.sheetnames:
        workbook.create_sheet('Employee List')
        workbook.save(get_schedule_file(year))
    sheet = workbook['Employee List']
    employees = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            employees.append(row[0])
    return employees

def save_employees(employees):
    year = datetime.datetime.now().year
    workbook = get_workbook(year)
    if 'Employee List' not in workbook.sheetnames:
        workbook.create_sheet('Employee List')
    sheet = workbook['Employee List']
    # Clear existing data
    sheet.delete_rows(2, sheet.max_row)
    # Write employees
    for idx, emp in enumerate(employees, start=2):
        sheet.cell(row=idx, column=1, value=emp)
    workbook.save(get_schedule_file(year))

def get_existing_schedule_years():
    years = set()
    for filename in os.listdir(SCHEDULING_FOLDER):
        if filename.endswith('On Call Scheduling.xlsx'):
            year_part = filename.split(' ')[0]
            if year_part.isdigit():
                years.add(int(year_part))
    return years

def load_previous_year_counts(year):
    # Load the workbook for the previous year
    prev_schedule_file = get_schedule_file(year)
    if not os.path.exists(prev_schedule_file):
        return {}, {}
    workbook = load_workbook(prev_schedule_file)
    # Load counts from the 'Reports' sheet
    if 'Reports' not in workbook.sheetnames:
        return {}, {}
    report_sheet = workbook['Reports']
    prev_primary_counts = {}
    prev_backup_counts = {}
    for row in report_sheet.iter_rows(min_row=2, values_only=True):
        emp, primary_count, backup_count = row
        prev_primary_counts[emp] = primary_count
        prev_backup_counts[emp] = backup_count
    return prev_primary_counts, prev_backup_counts

def load_previous_year_edit_differences(year):
    # Load original and edited counts from previous year's data
    original_counts = load_counts_from_sheet(year, 'Original Reports')
    edited_counts = load_counts_from_sheet(year, 'Reports')
    return original_counts, edited_counts

def load_counts_from_sheet(year, sheet_name):
    schedule_file = get_schedule_file(year)
    if not os.path.exists(schedule_file):
        return {}
    workbook = load_workbook(schedule_file)
    if sheet_name not in workbook.sheetnames:
        return {}
    sheet = workbook[sheet_name]
    counts = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        emp, primary_count, _ = row
        counts[emp] = primary_count
    return counts

def calculate_workload_differences(original_counts, edited_counts):
    differences = {}
    for emp in original_counts:
        original = original_counts.get(emp, 0)
        edited = edited_counts.get(emp, 0)
        differences[emp] = original - edited  # Positive if employee worked less
    return differences

def generate_schedule():
    clear_screen()
    employees = load_employees()

    if not employees:
        print_centered("Employee list is empty. Add employees first.", Fore.RED)
        pause()
        return

    # Prompt for the year
    while True:
        year_input = input("Enter the year for which you want to generate the schedule (e.g., 2025): ").strip()
        if year_input.isdigit() and int(year_input) >= datetime.datetime.now().year:
            year = int(year_input)
            break
        else:
            print_centered("Please enter a valid future year.", Fore.RED)

    # Check if schedule already exists
    existing_years = get_existing_schedule_years()
    if year in existing_years:
        confirm = input(f"A schedule for {year} already exists. Do you want to overwrite it? (y/n): ").strip().lower()
        if confirm != 'y':
            print_centered("Schedule generation canceled.", Fore.BLUE)
            pause()
            return

    # Load previous year's counts
    previous_year = year - 1
    prev_primary_counts, prev_backup_counts = load_previous_year_counts(previous_year)

    # Load differences due to schedule edits
    prev_original_counts, prev_edited_counts = load_previous_year_edit_differences(previous_year)
    workload_differences = calculate_workload_differences(prev_original_counts, prev_edited_counts)

    # Combine current and previous counts
    total_primary_counts = {emp: prev_primary_counts.get(emp, 0) for emp in employees}
    total_backup_counts = {emp: prev_backup_counts.get(emp, 0) for emp in employees}

    # Inform the user about workload differences
    if workload_differences:
        print_centered("Detected workload differences due to schedule edits in the previous year:", Fore.YELLOW)
        for emp, diff in workload_differences.items():
            if diff != 0:
                print_centered(f"{emp}: {'Worked less' if diff > 0 else 'Worked more'} by {abs(diff)} assignments.")
        adjust = input("Do you want to adjust the new schedule to compensate? (y/n): ").strip().lower()
        if adjust == 'y':
            # Adjust counts to compensate
            for emp in employees:
                total_primary_counts[emp] += workload_differences.get(emp, 0)
                # Ensure counts don't go negative
                total_primary_counts[emp] = max(total_primary_counts[emp], 0)

    # Proceed with schedule generation using 'total_primary_counts' and 'total_backup_counts'

    # Initialize counts
    primary_counts = {emp: 0 for emp in employees}
    backup_counts = {emp: 0 for emp in employees}

    # Create a list of all weeks in the year starting from the first Monday
    start_date = datetime.date(year, 1, 1)
    if start_date.weekday() != 0:
        start_date += datetime.timedelta(days=(7 - start_date.weekday()))
    weeks_in_year = 52

    week_dates = []
    for week in range(weeks_in_year):
        week_start = start_date + datetime.timedelta(weeks=week)
        week_dates.append(week_start)

    # Assign primaries evenly
    primary_schedule = {}
    employee_cycle = employees.copy()
    random.shuffle(employee_cycle)
    for week_start in week_dates:
        # Find employee with the least total primary assignments
        employee_cycle.sort(key=lambda emp: (total_primary_counts.get(emp, 0) + primary_counts[emp]))
        emp = employee_cycle[0]
        primary_schedule[week_start] = emp
        primary_counts[emp] += 1
        # Rotate the employee list to distribute assignments
        employee_cycle = employee_cycle[1:] + [emp]

    # Build date_primary mapping from primary_schedule
    date_primary = {}
    all_dates = []
    for week_start in week_dates:
        for i in range(7):
            date = week_start + datetime.timedelta(days=i)
            if date.year == year:
                date_primary[date] = primary_schedule[week_start]
                all_dates.append(date)

    # Initialize backup counts
    backup_counts = {emp: 0 for emp in employees}

    # Assign backups for each date
    date_backup1 = {}
    date_backup2 = {}
    last_backup1_date = {emp: datetime.date(year - 1, 12, 31) for emp in employees}
    last_backup2_date = {emp: datetime.date(year - 1, 12, 31) for emp in employees}

    for date in all_dates:
        primary = date_primary[date]

        # Exclude primary and apply constraints
        eligible_employees = [emp for emp in employees if emp != primary]

        # Remove employees who were backup the previous day
        previous_date = date - datetime.timedelta(days=1)
        if previous_date in date_backup1:
            prev_backup1 = date_backup1[previous_date]
            eligible_employees = [emp for emp in eligible_employees if emp != prev_backup1]
        if previous_date in date_backup2:
            prev_backup2 = date_backup2[previous_date]
            eligible_employees = [emp for emp in eligible_employees if emp != prev_backup2]

        # Remove employees who are primary the next day
        next_date = date + datetime.timedelta(days=1)
        if next_date in date_primary:
            next_primary = date_primary[next_date]
            eligible_employees = [emp for emp in eligible_employees if emp != next_primary]

        # Sort eligible employees by least total backup assignments
        eligible_employees.sort(key=lambda emp: (total_backup_counts.get(emp, 0) + backup_counts[emp]))

        # Assign Backup 1
        if eligible_employees:
            for emp in eligible_employees:
                if last_backup1_date[emp] + datetime.timedelta(days=1) < date:
                    date_backup1[date] = emp
                    backup_counts[emp] += 1
                    last_backup1_date[emp] = date
                    eligible_employees.remove(emp)
                    break
            else:
                # If no suitable candidate found, assign the next available
                emp = eligible_employees[0]
                date_backup1[date] = emp
                backup_counts[emp] += 1
                last_backup1_date[emp] = date
                eligible_employees.remove(emp)
        else:
            # If no eligible employees, assign any employee not primary
            emp = [e for e in employees if e != primary][0]
            date_backup1[date] = emp
            backup_counts[emp] += 1
            last_backup1_date[emp] = date

        # Assign Backup 2
        eligible_employees = [emp for emp in eligible_employees if emp != date_backup1[date]]
        if eligible_employees:
            for emp in eligible_employees:
                if last_backup2_date[emp] + datetime.timedelta(days=1) < date:
                    date_backup2[date] = emp
                    backup_counts[emp] += 1
                    last_backup2_date[emp] = date
                    break
            else:
                # If no suitable candidate found, assign the next available
                emp = eligible_employees[0]
                date_backup2[date] = emp
                backup_counts[emp] += 1
                last_backup2_date[emp] = date
        else:
            # If no eligible employees, assign any employee not primary or backup1
            emp = [e for e in employees if e not in [primary, date_backup1[date]]][0]
            date_backup2[date] = emp
            backup_counts[emp] += 1
            last_backup2_date[emp] = date

    # Create workbook and sheets
    workbook = get_workbook(year)

    # Clear existing month sheets
    for month in range(1, 13):
        month_name = datetime.date(year, month, 1).strftime('%B')
        if month_name in workbook.sheetnames:
            sheet = workbook[month_name]
            workbook.remove(sheet)
        workbook.create_sheet(month_name)

    # Populate calendar sheets
    create_calendar_sheets(workbook, date_primary, date_backup1, date_backup2, year)

    # Generate reports
    generate_reports(workbook, employees, primary_counts, backup_counts)
    # Save original counts to 'Original Reports' sheet
    generate_reports(workbook, employees, primary_counts, backup_counts, sheet_name='Original Reports')

    # Save the workbook
    workbook.save(get_schedule_file(year))
    print_centered(f"Schedule generated and saved to {get_schedule_file(year)}", Fore.BLUE)
    pause()

def create_calendar_sheets(workbook, date_primary, date_backup1, date_backup2, year):
    for month in range(1, 13):
        month_name = datetime.date(year, month, 1).strftime('%B')
        if month_name not in workbook.sheetnames:
            workbook.create_sheet(month_name)
        sheet = workbook[month_name]
        create_calendar_sheet(sheet, year, month, date_primary, date_backup1, date_backup2)

def create_calendar_sheet(sheet, year, month, date_primary, date_backup1, date_backup2):
    sheet.sheet_view.showGridLines = False  # Hide default gridlines

    # Styling
    day_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    font = Font(name='Calibri', size=11)
    align = Alignment(horizontal='center', vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Get first day of the month and number of days
    first_day = datetime.date(year, month, 1)
    if month == 12:
        num_days = 31
    else:
        next_month = datetime.date(year, month + 1, 1)
        num_days = (next_month - first_day).days
    start_day = first_day.weekday()  # Monday is 0

    # Add month and year title
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = first_day.strftime('%B %Y')
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.border = Border(bottom=Side(style='medium'))

    # Weekday headers
    weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for i, day in enumerate(weekdays, start=1):
        cell = sheet.cell(row=2, column=i)
        cell.value = day
        cell.font = Font(bold=True)
        cell.alignment = align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row = 3
    col = start_day + 1
    day_num = 1

    while day_num <= num_days:
        for col in range(col, 8):
            if day_num > num_days:
                break
            date = datetime.date(year, month, day_num)
            primary = date_primary.get(date, '')
            backup1 = date_backup1.get(date, '')
            backup2 = date_backup2.get(date, '')
            cell = sheet.cell(row=row, column=col)
            cell.value = f"Day {day_num}\nP:\n{primary}\n\nB1:\n{backup1}\n\nB2:\n{backup2}"
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
            cell.fill = day_fill
            cell.font = font
            cell.border = thin_border
            day_num += 1

        col = 1
        row += 1

    # Adjust the column widths based on the maximum content width
    for idx, col in enumerate(sheet.columns, 1):
        max_length = 0
        col_letter = get_column_letter(idx)  # Get the column letter using the index
        for cell in col:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if cell.value:
                cell_length = max(len(line) for line in str(cell.value).split('\n'))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Adjust row heights to add extra space between rows
    for row_idx in range(3, sheet.max_row + 1):
        row_height = sheet.row_dimensions[row_idx].height
        if not row_height:
            row_height = 15  # Default row height
        sheet.row_dimensions[row_idx].height = row_height + 7.5  # Add extra space (~10 pixels)

    # Freeze panes so that the weekday headers stay visible
    sheet.freeze_panes = 'A3'

def generate_reports(workbook, employees, primary_counts, backup_counts, sheet_name='Reports'):
    # Create or get the reports sheet
    if sheet_name in workbook.sheetnames:
        report_sheet = workbook[sheet_name]
        workbook.remove(report_sheet)
    report_sheet = workbook.create_sheet(sheet_name)

    # Write the report
    report_sheet.append(['Employee', 'Primary Count', 'Backup Count'])
    for emp in employees:
        report_sheet.append([emp, primary_counts.get(emp, 0), backup_counts.get(emp, 0)])

def recalculate_counts(employees, date_primary, date_backup1, date_backup2):
    primary_counts = {emp: 0 for emp in employees}
    backup_counts = {emp: 0 for emp in employees}

    #logging.debug('Starting to recalculate counts.')
    #logging.debug('Employees: %s', employees)

    for date in date_primary:
        primary = date_primary[date]
        if primary in primary_counts:
            primary_counts[primary] += 1
            #logging.debug('Incremented primary count for %s on %s. New count: %d', primary, date, primary_counts[primary])
        #else:
            #logging.warning('Primary employee %s on %s not found in employee list.', primary, date)

    for date in date_backup1:
        backup1 = date_backup1[date]
        if backup1 in backup_counts:
            backup_counts[backup1] += 1
            #logging.debug('Incremented backup1 count for %s on %s. New count: %d', backup1, date, backup_counts[backup1])
        #else:
            #logging.warning('Backup1 employee %s on %s not found in employee list.', backup1, date)

    for date in date_backup2:
        backup2 = date_backup2[date]
        if backup2 in backup_counts:
            backup_counts[backup2] += 1
            #logging.debug('Incremented backup2 count for %s on %s. New count: %d', backup2, date, backup_counts[backup2])
        #else:
            #logging.warning('Backup2 employee %s on %s not found in employee list.', backup2, date)

    #logging.debug('Finished recalculating counts.')
    #logging.debug('Final primary counts: %s', primary_counts)
    #logging.debug('Final backup counts: %s', backup_counts)

    return primary_counts, backup_counts


def load_schedule_data(year):
    workbook = get_workbook(year)
    date_primary = {}
    date_backup1 = {}
    date_backup2 = {}

    #logging.debug('Loading schedule data for year %d', year)

    for month in range(1, 13):
        month_name = datetime.date(year, month, 1).strftime('%B')
        if month_name in workbook.sheetnames:
            sheet = workbook[month_name]
            #logging.debug('Processing sheet: %s', month_name)
            for row in sheet.iter_rows(min_row=3):
                for cell in row:
                    if cell.value and not isinstance(cell, MergedCell):
                        content = cell.value
                        day_num = None
                        primary = ''
                        backup1 = ''
                        backup2 = ''
                        lines = content.split('\n')
                        i = 0
                        while i < len(lines):
                            line = lines[i]
                            if line.startswith('Day '):
                                day_num = int(line.strip().split()[1])
                                i += 1
                            elif line.strip() == 'P:':
                                if i + 1 < len(lines):
                                    primary = lines[i + 1].strip()
                                    i += 2
                                else:
                                    i += 1
                            elif line.strip() == 'B1:':
                                if i + 1 < len(lines):
                                    backup1 = lines[i + 1].strip()
                                    i += 2
                                else:
                                    i += 1
                            elif line.strip() == 'B2:':
                                if i + 1 < len(lines):
                                    backup2 = lines[i + 1].strip()
                                    i += 2
                                else:
                                    i += 1
                            else:
                                i += 1
                        if day_num:
                            date = datetime.date(year, month, day_num)
                            date_primary[date] = primary
                            date_backup1[date] = backup1
                            date_backup2[date] = backup2
                            #logging.debug('Loaded assignments for %s: Primary=%s, Backup1=%s, Backup2=%s', date, primary, backup1, backup2)
    return date_primary, date_backup1, date_backup2




def save_schedule_data(date_primary, date_backup1, date_backup2, year):
    workbook = get_workbook(year)
    #logging.debug('Saving schedule data for year %d', year)
    # Remove existing month sheets
    for month in range(1, 13):
        month_name = datetime.date(year, month, 1).strftime('%B')
        if month_name in workbook.sheetnames:
            del workbook[month_name]
            #logging.debug('Deleted sheet: %s', month_name)
    # Recreate month sheets
    for month in range(1, 13):
        month_name = datetime.date(year, month, 1).strftime('%B')
        workbook.create_sheet(month_name)
        #logging.debug('Created sheet: %s', month_name)
    create_calendar_sheets(workbook, date_primary, date_backup1, date_backup2, year)
    workbook.save(get_schedule_file(year))
    #logging.debug('Schedule data saved for year %d', year)



def manage_schedule_changes():
    clear_screen()
    print_centered("Manage Schedule Changes", Fore.CYAN)
    # Prompt for the year
    while True:
        year_input = input("Enter the year of the schedule you want to modify (e.g., 2025): ").strip()
        if year_input.isdigit():
            year = int(year_input)
            break
        else:
            print_centered("Please enter a valid year.", Fore.RED)
    workbook = get_workbook(year)
    date_primary, date_backup1, date_backup2 = load_schedule_data(year)
    if not date_primary:
        print_centered("No schedule data available. Generate schedule first.", Fore.RED)
        pause()
        return

    # Prompt for date or date range
    print()
    date_input = input("Enter date (YYYY-MM-DD) or date range (YYYY-MM-DD to YYYY-MM-DD): ").strip()
    date_format = "%Y-%m-%d"
    try:
        if 'to' in date_input:
            start_str, end_str = [s.strip() for s in date_input.split('to')]
            start_date = datetime.datetime.strptime(start_str, date_format).date()
            end_date = datetime.datetime.strptime(end_str, date_format).date()
        else:
            start_date = end_date = datetime.datetime.strptime(date_input, date_format).date()
    except ValueError:
        print_centered("Invalid date format.", Fore.RED)
        pause()
        return

    # Collect dates within the range
    delta = datetime.timedelta(days=1)
    dates_to_modify = []
    current_date = start_date
    while current_date <= end_date:
        if current_date in date_primary:
            dates_to_modify.append(current_date)
        current_date += delta

    if not dates_to_modify:
        print_centered("No schedule entries found for the specified date(s).", Fore.RED)
        pause()
        return

    employees = load_employees()
    employee_dict = {str(idx): emp for idx, emp in enumerate(employees, 1)}

    for date in dates_to_modify:
        clear_screen()
        print_centered(f"Date: {date.strftime('%Y-%m-%d')}", Fore.YELLOW)
        print()
        print_centered(f"1. Primary: {date_primary[date]}")
        print_centered(f"2. Backup 1: {date_backup1[date]}")
        print_centered(f"3. Backup 2: {date_backup2[date]}")
        print_centered("4. Skip to next date")
        print()
        choice = input("Enter the number of the assignment you want to change: ").strip()
        if choice not in ['1', '2', '3']:
            continue  # Skip to next date
        role = ''
        if choice == '1':
            role = 'Primary'
            current_employee = date_primary[date]
        elif choice == '2':
            role = 'Backup 1'
            current_employee = date_backup1[date]
        elif choice == '3':
            role = 'Backup 2'
            current_employee = date_backup2[date]
        else:
            continue

        print()
        print_centered(f"Current {role}: {current_employee}", Fore.YELLOW)
        print_centered("Available Employees:", Fore.CYAN)
        for idx, emp in employee_dict.items():
            print(center_text(f"{idx}. {emp}"))

        new_emp_idx = input("Enter the number of the employee to assign: ").strip()
        if new_emp_idx not in employee_dict:
            print_centered("Invalid selection.", Fore.RED)
            pause()
            continue
        new_employee = employee_dict[new_emp_idx]

        # Update assignment
        if role == 'Primary':
            date_primary[date] = new_employee
        elif role == 'Backup 1':
            date_backup1[date] = new_employee
        elif role == 'Backup 2':
            date_backup2[date] = new_employee

        #logging.info('Changed %s on %s from %s to %s', role, date, current_employee, new_employee)

        print_centered(f"{role} for {date.strftime('%Y-%m-%d')} updated to {new_employee}", Fore.GREEN)
        pause()

    # Save updated schedule
    save_schedule_data(date_primary, date_backup1, date_backup2, year)

    # Recalculate counts and update reports
    employees = load_employees()
    primary_counts, backup_counts = recalculate_counts(employees, date_primary, date_backup1, date_backup2)
    workbook = get_workbook(year)
    generate_reports(workbook, employees, primary_counts, backup_counts)
    workbook.save(get_schedule_file(year))

    print_centered("Schedule changes saved and reports updated.", Fore.BLUE)
    pause()


def manage_employees():
    clear_screen()
    employees = load_employees()
    while True:
        clear_screen()
        print_centered("Manage Employees", Fore.CYAN)
        print()
        print_centered("Current Employee List:", Fore.YELLOW)
        for emp in employees:
            print(center_text(emp))
        print()
        print_centered("Options:", Fore.CYAN)
        print_centered("1. Add Employee")
        print_centered("2. Remove Employee")
        print_centered("3. Back to Main Menu")
        print()
        choice = input("Enter your choice: ").strip()
        if choice == '1':
            name = input("Enter employee name: ").strip()
            if name and name != 'devintest' and name not in employees:
                employees.append(name)
                save_employees(employees)
            else:
                print_centered("Invalid name or already exists.", Fore.RED)
                pause()
        elif choice == '2':
            name = input("Enter employee name to remove: ").strip()
            if name in employees:
                employees.remove(name)
                save_employees(employees)
            else:
                print_centered("Employee not found.", Fore.RED)
                pause()
        elif choice == '3':
            break
        else:
            print_centered("Invalid choice.", Fore.RED)
            pause()

def view_reports():
    clear_screen()
    # Prompt for the year
    while True:
        year_input = input("Enter the year of the report you want to view (e.g., 2025): ").strip()
        if year_input.isdigit():
            year = int(year_input)
            break
        else:
            print_centered("Please enter a valid year.", Fore.RED)
    workbook = get_workbook(year)
    if 'Reports' not in workbook.sheetnames:
        print_centered("No report data available. Generate schedule first.", Fore.RED)
        pause()
        return
    report_sheet = workbook['Reports']
    employees = []
    primary_counts = {}
    backup_counts = {}

    for row in report_sheet.iter_rows(min_row=2, values_only=True):
        emp, primary_count, backup_count = row
        employees.append(emp)
        primary_counts[emp] = primary_count
        backup_counts[emp] = backup_count

    clear_screen()
    print_centered(f"Employee On-Call Report for {year}", Fore.CYAN)
    print()
    header = f"{'Employee':<20}{'Primary Count':<15}{'Backup Count'}"
    print_centered(header, Fore.YELLOW)
    for emp in employees:
        line = f"{emp:<20}{primary_counts.get(emp,0):<15}{backup_counts.get(emp,0)}"
        print(center_text(line))
    pause()


def main_menu():
    response = ""
    previous_responses = []
    while True:
        clear_screen()
        for res in previous_responses[-5:]:
            print(res)
        print(Back.LIGHTBLACK_EX + ' ' * os.get_terminal_size().columns)
        print_centered("On Call Scheduling Program", Fore.GREEN + Style.BRIGHT)
        print()
        print_centered("1. Generate Schedule")
        print_centered("2. View/Edit Employees")
        print_centered("3. View Reports")
        print_centered("4. Manage Schedule Changes")
        print_centered("5. Exit")
        print()
        if response:
            print(response)
            previous_responses.append(response)
            response = ""
        print()
        choice = input("Enter your choice: ").strip()
        if choice == '1':
            generate_schedule()
            response = "Schedule generated successfully."
        elif choice == '2':
            manage_employees()
            response = "Employee list updated."
        elif choice == '3':
            view_reports()
            response = ""
        elif choice == '4':
            manage_schedule_changes()
            response = ""
        elif choice == '5':
            clear_screen()
            sys.exit()
        elif choice == 'devintest':
            employees = DEVINTEST_EMPLOYEES.copy()
            save_employees(employees)
            response = "Devintest mode activated with test employees."
        else:
            response = "Invalid choice. Please try again."

if __name__ == "__main__":
    main_menu()
